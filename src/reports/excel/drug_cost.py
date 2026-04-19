"""
Phase 3c: Drug Cost Analysis Excel Report Builder.

Generates output/excel/drug_cost.xlsx with three sheets:
  - Summary: Brand vs. Generic spend, Top 10 drugs, Therapeutic class breakdown
  - Detail: full Paid claim-level data with filters + frozen header
  - Top Drugs Chart: horizontal bar chart of top 10 drugs by gross cost
"""

from __future__ import annotations

import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.utils import get_column_letter

from src.config import OUTPUT_DIR
from src.ingestion import run_query
from src.dashboard.routes._filters import _inject_filter
from src.reports.excel._utils import (
    _bold_cell,
    _get_date_range_label,
    _header_row,
    _load_queries,
    _set_col_widths,
)
from src.reports.excel.constants import (
    CHART_HEIGHT_LG,
    CHART_STYLE,
    CHART_WIDTH_LG,
    COLOR_AMBER,
    FMT_CURRENCY,
)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_SQL_FILE = (
    Path(__file__).resolve().parent.parent.parent / "sql" / "drug_cost.sql"
)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary(wb: Workbook, queries: dict[str, str], extra_where: str = "") -> pd.DataFrame:
    """Build the Summary sheet; returns top_10_drugs df for the chart sheet."""
    ws = wb.create_sheet("Summary")

    # Run queries (apply optional filter from dashboard)
    df_brand_generic = run_query(_inject_filter(queries["brand_vs_generic"], extra_where))
    df_top10 = run_query(_inject_filter(queries["top_10_drugs"], extra_where))
    df_tc = run_query(_inject_filter(queries["therapeutic_class_spend"], extra_where))

    today_str = datetime.date.today().strftime("%Y-%m-%d")

    row = 1

    # ---- Title block --------------------------------------------------------
    title_cell = ws.cell(row=row, column=1, value="Drug Cost Analysis Report")
    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    title_cell.alignment = Alignment(horizontal="center")
    row += 1

    # Subtitle
    ws.cell(row=row, column=1, value=f"Generated: {today_str}")
    row += 1

    # Blank separator
    row += 1

    # ---- Brand vs. Generic section ------------------------------------------
    _bold_cell(ws, row, 1, "Brand vs. Generic Spend", size=12)
    row += 1

    bvg_headers = [
        "Drug Type", "Claims", "Total Gross Cost", "Avg Cost per Claim", "% of Total Cost"
    ]
    _header_row(ws, row, bvg_headers)
    row += 1

    for _, brow in df_brand_generic.iterrows():
        ws.cell(row=row, column=1, value=str(brow["drug_type"]))
        ws.cell(row=row, column=2, value=int(brow["claim_count"]))

        cell_total = ws.cell(row=row, column=3, value=float(brow["total_gross_cost"]))
        cell_total.number_format = FMT_CURRENCY

        cell_avg = ws.cell(row=row, column=4, value=float(brow["avg_gross_cost"]))
        cell_avg.number_format = FMT_CURRENCY

        cell_pct = ws.cell(row=row, column=5, value=float(brow["pct_of_total_cost"]))
        cell_pct.number_format = '0.00"%"'

        row += 1

    # Blank separator
    row += 1

    # ---- Top 10 Drugs section -----------------------------------------------
    _bold_cell(ws, row, 1, "Top 10 Drugs by Cost", size=12)
    row += 1

    top10_headers = [
        "Brand Name", "Generic Name", "Type", "Therapeutic Class",
        "Claims", "Gross Cost", "Total Paid"
    ]
    _header_row(ws, row, top10_headers)
    row += 1

    amber_fill = PatternFill(start_color=COLOR_AMBER, end_color=COLOR_AMBER, fill_type="solid")

    for i, (_, trow) in enumerate(df_top10.iterrows()):
        ws.cell(row=row, column=1, value=str(trow["brand_name"]))
        ws.cell(row=row, column=2, value=str(trow["generic_name"]))
        ws.cell(row=row, column=3, value=str(trow["drug_type"]))
        ws.cell(row=row, column=4, value=str(trow["therapeutic_class"]))
        ws.cell(row=row, column=5, value=int(trow["claim_count"]))

        cell_gross = ws.cell(row=row, column=6, value=float(trow["total_gross_cost"]))
        cell_gross.number_format = FMT_CURRENCY

        cell_paid = ws.cell(row=row, column=7, value=float(trow["total_paid_amount"]))
        cell_paid.number_format = FMT_CURRENCY

        # Highlight the top row (highest cost drug) with light amber fill
        if i == 0:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = amber_fill

        row += 1

    # Blank separator
    row += 1

    # ---- Therapeutic Class section ------------------------------------------
    _bold_cell(ws, row, 1, "Spend by Therapeutic Class", size=12)
    row += 1

    tc_headers = [
        "Therapeutic Class", "Claims", "Total Gross Cost",
        "Avg Gross Cost", "Brand Cost", "Generic Cost"
    ]
    _header_row(ws, row, tc_headers)
    row += 1

    tc_data_start = row
    for _, tcrow in df_tc.iterrows():
        ws.cell(row=row, column=1, value=str(tcrow["therapeutic_class"]))
        ws.cell(row=row, column=2, value=int(tcrow["claim_count"]))

        cell_total = ws.cell(row=row, column=3, value=float(tcrow["total_gross_cost"]))
        cell_total.number_format = FMT_CURRENCY

        cell_avg = ws.cell(row=row, column=4, value=float(tcrow["avg_gross_cost"]))
        cell_avg.number_format = FMT_CURRENCY

        cell_brand = ws.cell(row=row, column=5, value=float(tcrow["brand_cost"]))
        cell_brand.number_format = FMT_CURRENCY

        cell_generic = ws.cell(row=row, column=6, value=float(tcrow["generic_cost"]))
        cell_generic.number_format = FMT_CURRENCY

        row += 1

    # TOTAL row for Therapeutic Class
    total_font = Font(bold=True)
    ws.cell(row=row, column=1, value="TOTAL").font = total_font
    ws.cell(row=row, column=2,
            value=f"=SUM(B{tc_data_start}:B{row - 1})").font = total_font
    tc_gross = ws.cell(row=row, column=3,
                       value=f"=SUM(C{tc_data_start}:C{row - 1})")
    tc_gross.font = total_font
    tc_gross.number_format = FMT_CURRENCY
    tc_brand = ws.cell(row=row, column=5,
                       value=f"=SUM(E{tc_data_start}:E{row - 1})")
    tc_brand.font = total_font
    tc_brand.number_format = FMT_CURRENCY
    tc_generic = ws.cell(row=row, column=6,
                         value=f"=SUM(F{tc_data_start}:F{row - 1})")
    tc_generic.font = total_font
    tc_generic.number_format = FMT_CURRENCY
    row += 1

    # Column widths
    _set_col_widths(ws, [24, 24, 14, 22, 12, 16, 16])

    return df_top10


def _build_detail(wb: Workbook, extra_where: str = "") -> None:
    """Build the Detail sheet with full Paid claim-level data."""
    ws = wb.create_sheet("Detail")

    base_q = "SELECT * FROM claims WHERE claim_status = 'Paid' ORDER BY gross_cost DESC"
    df = run_query(_inject_filter(base_q, extra_where))

    headers = df.columns.tolist()
    _header_row(ws, 1, headers)

    date_fmt = "YYYY-MM-DD"
    currency_fmt = '"$"#,##0.00'
    date_cols = {"service_date", "paid_date"}
    cost_cols = {"gross_cost", "member_copay", "plan_paid", "total_paid"}
    numeric_cols = cost_cols | {"quantity", "days_supply", "formulary_tier"}

    for r_idx, (_, data_row) in enumerate(df.iterrows(), start=2):
        for c_idx, col_name in enumerate(headers, start=1):
            raw_val = data_row[col_name]

            if pd.isna(raw_val):
                val = None
            elif col_name in date_cols:
                try:
                    val = datetime.datetime.strptime(str(raw_val), "%Y-%m-%d").date()
                except ValueError:
                    val = str(raw_val)
            elif hasattr(raw_val, "item"):
                val = raw_val.item()
            else:
                val = raw_val

            cell = ws.cell(row=r_idx, column=c_idx, value=val)

            if col_name in date_cols:
                cell.number_format = date_fmt
            elif col_name in cost_cols:
                cell.number_format = currency_fmt

            if col_name in numeric_cols:
                cell.alignment = Alignment(horizontal="right")

    # AutoFilter covering header + all data rows
    row_count = len(df)
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{row_count + 1}"

    # Freeze top row
    ws.freeze_panes = "A2"

    # Set reasonable column widths
    col_widths = {
        "claim_id": 14, "member_id": 12, "member_name": 20, "plan_id": 10,
        "plan_name": 22, "service_date": 14, "paid_date": 14, "ndc": 14,
        "brand_name": 22, "generic_name": 22, "drug_type": 12, "therapeutic_class": 22,
        "formulary_tier": 14, "strength": 10, "days_supply": 12, "quantity": 10,
        "pharmacy_id": 14, "pharmacy_name": 22, "pharmacy_city": 16,
        "pharmacy_state": 14, "gross_cost": 14, "member_copay": 14,
        "plan_paid": 14, "total_paid": 14, "claim_status": 14,
    }
    for c_idx, col_name in enumerate(headers, start=1):
        width = col_widths.get(col_name, 15)
        ws.column_dimensions[get_column_letter(c_idx)].width = width


def _build_chart_sheet(wb: Workbook, df_top10: pd.DataFrame, extra_where: str = "") -> None:
    """Build the Top Drugs Chart sheet with a horizontal bar chart."""
    ws = wb.create_sheet("Top Drugs Chart")

    # Write the data table that the chart references
    ws.cell(row=1, column=1, value="Brand Name").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Gross Cost").font = Font(bold=True)

    for r_idx, (_, trow) in enumerate(df_top10.iterrows(), start=2):
        ws.cell(row=r_idx, column=1, value=str(trow["brand_name"]))
        ws.cell(row=r_idx, column=2, value=float(trow["total_gross_cost"]))

    n_drugs = len(df_top10)

    date_range = _get_date_range_label(extra_where)

    # Build horizontal bar chart (direction="bar" makes it horizontal)
    chart = BarChart()
    chart.type = "bar"   # "bar" = horizontal bars; "col" = vertical columns
    chart.title = f"Top 10 Drugs by Gross Cost (Paid Claims){date_range}"
    chart.x_axis.title = "Gross Cost"
    chart.y_axis.title = "Drug"
    chart.style = CHART_STYLE
    chart.width = CHART_WIDTH_LG
    chart.height = CHART_HEIGHT_LG
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=n_drugs + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=n_drugs + 1)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Place chart below the data table
    anchor_row = n_drugs + 3
    ws.add_chart(chart, f"A{anchor_row}")

    # Column widths for the data table
    _set_col_widths(ws, [28, 16])


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_drug_report(extra_where: str = "") -> str:
    """Build the Drug Cost Analysis Excel report. Returns the output file path."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    queries = _load_queries(_SQL_FILE)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Summary
    df_top10 = _build_summary(wb, queries, extra_where)

    # Sheet 2: Detail
    _build_detail(wb, extra_where)

    # Sheet 3: Top Drugs Chart
    _build_chart_sheet(wb, df_top10, extra_where)

    output_path = OUTPUT_DIR / "drug_cost.xlsx"
    wb.save(str(output_path))

    return str(output_path)
