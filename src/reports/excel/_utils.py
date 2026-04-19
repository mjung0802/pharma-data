"""
Shared helper utilities for Excel report builders.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl.cell import Cell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet


def _load_queries(sql_file: Path) -> dict[str, str]:
    """Parse a SQL file into a dict of {query_name: sql_body}.

    The file is split on ``-- QUERY:`` markers; the text before the first
    marker is treated as a preamble and discarded.
    """
    sql_text = sql_file.read_text(encoding="utf-8")
    sections = sql_text.split("-- QUERY:")
    queries: dict[str, str] = {}
    for section in sections[1:]:  # skip preamble before first marker
        lines = section.strip().splitlines()
        name = lines[0].strip()
        body = "\n".join(lines[1:]).strip()
        queries[name] = body
    return queries


def _set_col_widths(ws: Worksheet, widths: list[int]) -> None:
    """Set column widths for the first len(widths) columns."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _bold_cell(ws: Worksheet, row: int, col: int, value, size: int = 11) -> Cell:
    """Write *value* into (row, col) with a bold font; return the cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, size=size)
    return cell


def _header_row(ws: Worksheet, row: int, headers: list[str], bold: bool = True) -> None:
    """Write *headers* across columns starting at column 1 in *row*."""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        if bold:
            cell.font = Font(bold=True)


def _apply_print_settings(ws: Worksheet) -> None:
    """Apply Letter/landscape print settings with standard margins."""
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER      # 1 = US Letter
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.75, bottom=0.75, header=0.3, footer=0.3
    )
    if ws.max_row and ws.max_column:
        last_col = get_column_letter(ws.max_column)
        ws.print_area = f"A1:{last_col}{ws.max_row}"


def _get_date_range_label(extra_where: str) -> str:
    """Return ' (YYYY-MM-DD to YYYY-MM-DD)' derived from data, or ''."""
    from src.ingestion import run_query
    from src.dashboard.routes._filters import _inject_filter
    sql = "SELECT MIN(service_date) AS min_d, MAX(service_date) AS max_d FROM claims"
    df = run_query(_inject_filter(sql, extra_where))
    if df.empty or df.iloc[0]["min_d"] is None:
        return ""
    return f" ({df.iloc[0]['min_d']} to {df.iloc[0]['max_d']})"
