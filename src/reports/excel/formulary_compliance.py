"""
Phase 3d: Formulary & Tier Compliance Excel Report Builder.

Generates output/excel/formulary_compliance.xlsx with three sheets:
  - Summary: KPIs, tier distribution, cost by tier and drug type
  - Detail: full Paid claim-level data with filters + frozen header
  - Tier Chart: vertical bar chart of gross cost by formulary tier
"""

from __future__ import annotations

import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.utils import get_column_letter

from src.config import OUTPUT_DIR
from src.ingestion import run_query
from src.dashboard.routes._filters import _inject_filter
from src.reports.excel._utils import (
    _bold_cell,
    _get_date_range_label,
    _header_row,
    _load_queries,
    _set_col_widths,
)
from src.reports.excel.constants import (
    CHART_HEIGHT_MD,
    CHART_STYLE,
    CHART_WIDTH_SM,
    COLOR_RED_LIGHT,
    FMT_CURRENCY,
)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_SQL_FILE = (
    Path(__file__).resolve().parent.parent.parent / "sql" / "formulary_compliance.sql"
)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary(wb: Workbook, queries: dict[str, str], extra_where: str = "") -> pd.DataFrame:
    """Build the Summary sheet; returns tier_distribution df for the chart sheet."""
    ws = wb.create_sheet("Summary")

    # Run queries (apply optional filter from dashboard)
    df_tier = run_query(_inject_filter(queries["tier_distribution"], extra_where))
    df_generic = run_query(_inject_filter(queries["generic_fill_rate"], extra_where))
    df_formulary = run_query(_inject_filter(queries["on_formulary_rate"], extra_where))
    df_cost_tier = run_query(_inject_filter(queries["cost_per_tier"], extra_where))

    today_str = datetime.date.today().strftime("%Y-%m-%d")

    row = 1

    # ---- Title block --------------------------------------------------------
    title_cell = ws.cell(row=row, column=1, value="Formulary & Tier Compliance Report")
    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    title_cell.alignment = Alignment(horizontal="center")
    row += 1

    # Subtitle
    ws.cell(row=row, column=1, value=f"Generated: {today_str}")
    row += 1

    # Blank separator
    row += 1

    # ---- Key Metrics section -------------------------------------------------
    # Formulas reference the Detail sheet (Paid claims only).
    # Detail columns: A=claim_id … K=drug_type, M=formulary_tier
    _bold_cell(ws, row, 1, "Key Metrics", size=12)
    row += 1

    _header_row(ws, row, ["Metric", "Value"])
    row += 1

    kpis: list[tuple[str, object, str | None]] = [
        ("Generic Fill Rate",
         "=IF(COUNTA(Detail!A2:A10000)=0,0,"
         "COUNTIF(Detail!K2:K10000,\"Generic\")/COUNTA(Detail!A2:A10000)*100)",
         '0.00"%"'),
        ("On-Formulary Rate",
         "=IF(COUNTA(Detail!A2:A10000)=0,0,"
         "(COUNTIF(Detail!M2:M10000,1)+COUNTIF(Detail!M2:M10000,2))"
         "/COUNTA(Detail!A2:A10000)*100)",
         '0.00"%"'),
        ("Total Paid Claims", "=COUNTA(Detail!A2:A10000)",                  None),
        ("Generic Claims",    '=COUNTIF(Detail!K2:K10000,"Generic")',        None),
        ("Brand Claims",      '=COUNTIF(Detail!K2:K10000,"Brand")',          None),
    ]
    for metric, value, fmt in kpis:
        ws.cell(row=row, column=1, value=metric)
        cell = ws.cell(row=row, column=2, value=value)
        if fmt is not None:
            cell.number_format = fmt
        row += 1

    # Blank separator
    row += 1

    # ---- Tier Distribution section ------------------------------------------
    _bold_cell(ws, row, 1, "Tier Distribution", size=12)
    row += 1

    tier_headers = [
        "Formulary Tier", "Claims", "Total Gross Cost", "Avg Gross Cost", "Avg Member Copay"
    ]
    _header_row(ws, row, tier_headers)
    row += 1

    red_fill = PatternFill(start_color=COLOR_RED_LIGHT, end_color=COLOR_RED_LIGHT, fill_type="solid")

    tier_data_start = row
    for _, trow in df_tier.iterrows():
        tier_val = int(trow["formulary_tier"])
        ws.cell(row=row, column=1, value=tier_val)
        ws.cell(row=row, column=2, value=int(trow["claim_count"]))

        cell_total = ws.cell(row=row, column=3, value=float(trow["total_gross_cost"]))
        cell_total.number_format = FMT_CURRENCY

        cell_avg = ws.cell(row=row, column=4, value=float(trow["avg_gross_cost"]))
        cell_avg.number_format = FMT_CURRENCY

        cell_copay = ws.cell(row=row, column=5, value=float(trow["avg_member_copay"]))
        cell_copay.number_format = FMT_CURRENCY

        # Highlight Tier 3 rows with light red fill
        if tier_val == 3:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = red_fill

        row += 1

    # TOTAL row for Tier Distribution
    total_font = Font(bold=True)
    ws.cell(row=row, column=1, value="TOTAL").font = total_font
    ws.cell(row=row, column=2,
            value=f"=SUM(B{tier_data_start}:B{row - 1})").font = total_font
    tier_total_cost = ws.cell(row=row, column=3,
                              value=f"=SUM(C{tier_data_start}:C{row - 1})")
    tier_total_cost.font = total_font
    tier_total_cost.number_format = '"$"#,##0.00'
    row += 1

    # Blank separator
    row += 1

    # ---- Cost by Tier and Drug Type section ---------------------------------
    _bold_cell(ws, row, 1, "Cost by Tier and Drug Type", size=12)
    row += 1

    cost_headers = [
        "Formulary Tier", "Drug Type", "Claims", "Total Gross Cost", "Avg Member Copay"
    ]
    _header_row(ws, row, cost_headers)
    row += 1

    for _, crow in df_cost_tier.iterrows():
        ws.cell(row=row, column=1, value=int(crow["formulary_tier"]))
        ws.cell(row=row, column=2, value=str(crow["drug_type"]))
        ws.cell(row=row, column=3, value=int(crow["claim_count"]))

        cell_total = ws.cell(row=row, column=4, value=float(crow["total_gross_cost"]))
        cell_total.number_format = '"$"#,##0.00'

        cell_copay = ws.cell(row=row, column=5, value=float(crow["avg_member_copay"]))
        cell_copay.number_format = '"$"#,##0.00'

        row += 1

    # Column widths
    _set_col_widths(ws, [20, 14, 18, 18, 18])

    return df_tier


def _build_detail(wb: Workbook, extra_where: str = "") -> None:
    """Build the Detail sheet with full Paid claim-level data."""
    ws = wb.create_sheet("Detail")

    base_q = "SELECT * FROM claims WHERE claim_status = 'Paid' ORDER BY formulary_tier ASC, gross_cost DESC"
    df = run_query(_inject_filter(base_q, extra_where))

    headers = df.columns.tolist()
    _header_row(ws, 1, headers)

    date_fmt = "YYYY-MM-DD"
    currency_fmt = '"$"#,##0.00'
    date_cols = {"service_date", "paid_date"}
    cost_cols = {"gross_cost", "member_copay", "plan_paid", "total_paid"}
    numeric_cols = cost_cols | {"quantity", "days_supply", "formulary_tier"}

    for r_idx, (_, data_row) in enumerate(df.iterrows(), start=2):
        for c_idx, col_name in enumerate(headers, start=1):
            raw_val = data_row[col_name]

            if pd.isna(raw_val):
                val = None
            elif col_name in date_cols:
                try:
                    val = datetime.datetime.strptime(str(raw_val), "%Y-%m-%d").date()
                except ValueError:
                    val = str(raw_val)
            elif hasattr(raw_val, "item"):
                val = raw_val.item()
            else:
                val = raw_val

            cell = ws.cell(row=r_idx, column=c_idx, value=val)

            if col_name in date_cols:
                cell.number_format = date_fmt
            elif col_name in cost_cols:
                cell.number_format = currency_fmt

            if col_name in numeric_cols:
                cell.alignment = Alignment(horizontal="right")

    # AutoFilter covering header + all data rows
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"

    # Freeze top row
    ws.freeze_panes = "A2"

    # Set reasonable column widths
    col_widths = {
        "claim_id": 14, "member_id": 12, "member_name": 20, "plan_id": 10,
        "plan_name": 22, "service_date": 14, "paid_date": 14, "ndc": 14,
        "brand_name": 22, "generic_name": 22, "drug_type": 12, "therapeutic_class": 22,
        "formulary_tier": 14, "strength": 10, "days_supply": 12, "quantity": 10,
        "pharmacy_id": 14, "pharmacy_name": 22, "pharmacy_city": 16,
        "pharmacy_state": 14, "gross_cost": 14, "member_copay": 14,
        "plan_paid": 14, "total_paid": 14, "claim_status": 14,
    }
    for c_idx, col_name in enumerate(headers, start=1):
        width = col_widths.get(col_name, 15)
        ws.column_dimensions[get_column_letter(c_idx)].width = width


def _build_chart_sheet(wb: Workbook, df_tier: pd.DataFrame, extra_where: str = "") -> None:
    """Build the Tier Chart sheet with a vertical bar chart of gross cost by tier."""
    ws = wb.create_sheet("Tier Chart")

    # Write the data table that the chart references
    ws.cell(row=1, column=1, value="Formulary Tier").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Total Gross Cost").font = Font(bold=True)

    for r_idx, (_, trow) in enumerate(df_tier.iterrows(), start=2):
        ws.cell(row=r_idx, column=1, value=int(trow["formulary_tier"]))
        ws.cell(row=r_idx, column=2, value=float(trow["total_gross_cost"]))

    n_tiers = len(df_tier)

    date_range = _get_date_range_label(extra_where)

    # Build vertical column chart
    chart = BarChart()
    chart.type = "col"   # "col" = vertical columns
    chart.title = f"Gross Cost by Formulary Tier (Paid Claims){date_range}"
    chart.y_axis.title = "Total Gross Cost"
    chart.x_axis.title = "Formulary Tier"
    chart.style = CHART_STYLE
    chart.width = CHART_WIDTH_SM
    chart.height = CHART_HEIGHT_MD
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=n_tiers + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=n_tiers + 1)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Place chart below the data table
    anchor_row = n_tiers + 3
    ws.add_chart(chart, f"A{anchor_row}")

    # Column widths for the data table
    _set_col_widths(ws, [16, 18])


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_formulary_report(extra_where: str = "") -> str:
    """Build the Formulary & Tier Compliance Excel report. Returns the output file path."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    queries = _load_queries(_SQL_FILE)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Summary
    df_tier = _build_summary(wb, queries, extra_where)

    # Sheet 2: Detail
    _build_detail(wb, extra_where)

    # Sheet 3: Tier Chart
    _build_chart_sheet(wb, df_tier, extra_where)

    output_path = OUTPUT_DIR / "formulary_compliance.xlsx"
    wb.save(str(output_path))

    return str(output_path)
