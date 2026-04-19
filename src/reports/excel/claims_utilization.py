"""
Phase 3b: Claims Utilization Excel Report Builder.

Generates output/excel/claims_utilization.xlsx with three sheets:
  - Summary: KPIs, status breakdown, monthly trend
  - Detail: full claim-level data with filters + frozen header
  - Monthly Trend Chart: bar chart of monthly claim volume
"""

from __future__ import annotations

import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.utils import get_column_letter

from src.config import OUTPUT_DIR
from src.ingestion import run_query
from src.dashboard.routes._filters import _inject_filter
from src.reports.excel._utils import (
    _apply_print_settings,
    _bold_cell,
    _get_date_range_label,
    _header_row,
    _load_queries,
    _set_col_widths,
)
from src.reports.excel.constants import (
    CHART_HEIGHT_MD,
    CHART_STYLE,
    CHART_WIDTH_MD,
    COLOR_RED_LIGHT,
    FMT_CURRENCY,
)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_SQL_FILE = (
    Path(__file__).resolve().parent.parent.parent / "sql" / "claims_utilization.sql"
)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary(wb: Workbook, queries: dict[str, str], extra_where: str = "") -> pd.DataFrame:
    """Build the Summary sheet; returns monthly_trend df for the chart sheet."""
    ws = wb.create_sheet("Summary")

    # Run queries (apply optional filter from dashboard)
    df_status = run_query(_inject_filter(queries["status_summary"], extra_where))
    df_monthly = run_query(_inject_filter(queries["monthly_trend"], extra_where))
    df_turnaround = run_query(_inject_filter(queries["turnaround_stats"], extra_where))

    today_str = datetime.date.today().strftime("%Y-%m-%d")

    row = 1

    # ---- Title block --------------------------------------------------------
    # Title row (merged)
    title_cell = ws.cell(row=row, column=1, value="Claims Utilization Report")
    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    title_cell.alignment = Alignment(horizontal="center")
    row += 1

    # Subtitle
    ws.cell(row=row, column=1, value=f"Generated: {today_str}")
    row += 1

    # Blank separator
    row += 1

    # ---- KPI section --------------------------------------------------------
    # Formulas reference the Detail sheet so any analyst can inspect the logic.
    # Detail columns: A=claim_id … U=gross_cost, X=total_paid, Y=claim_status
    _bold_cell(ws, row, 1, "Key Metrics", size=12)
    row += 1

    _header_row(ws, row, ["Metric", "Value"])
    row += 1

    kpis: list[tuple[str, object, str | None]] = [
        ("Overall Paid Rate",
         "=IF(COUNTA(Detail!A2:A10000)=0,0,"
         "COUNTIF(Detail!Y2:Y10000,\"Paid\")/COUNTA(Detail!A2:A10000)*100)",
         '0.00"%"'),
        ("Total Claims",    "=COUNTA(Detail!A2:A10000)",                   None),
        ("Paid Claims",     '=COUNTIF(Detail!Y2:Y10000,"Paid")',            None),
        ("Total Gross Cost","=SUM(Detail!U2:U10000)",                       FMT_CURRENCY),
        ("Total Paid (Plan)","=SUM(Detail!X2:X10000)",                      FMT_CURRENCY),
    ]
    for metric, value, fmt in kpis:
        ws.cell(row=row, column=1, value=metric)
        cell = ws.cell(row=row, column=2, value=value)
        if fmt is not None:
            cell.number_format = fmt
        row += 1

    # Blank separator
    row += 1

    # ---- Status Breakdown ---------------------------------------------------
    _bold_cell(ws, row, 1, "Claim Status Breakdown", size=12)
    row += 1

    status_headers = ["Claim Status", "Count", "Total Paid", "Paid Rate (%)"]
    _header_row(ws, row, status_headers)
    row += 1

    red_fill = PatternFill(start_color=COLOR_RED_LIGHT, end_color=COLOR_RED_LIGHT, fill_type="solid")

    status_data_start = row
    for _, srow in df_status.iterrows():
        ws.cell(row=row, column=1, value=srow["claim_status"])
        ws.cell(row=row, column=2, value=int(srow["claim_count"]))
        ws.cell(row=row, column=3, value=float(srow["total_paid"]))
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=float(srow["paid_rate"]))

        # Highlight rejected / reversed rows
        if srow["claim_status"] in ("Rejected", "Reversed"):
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = red_fill
        row += 1

    # TOTAL row for Status Breakdown
    total_font = Font(bold=True)
    ws.cell(row=row, column=1, value="TOTAL").font = total_font
    ws.cell(row=row, column=2,
            value=f"=SUM(B{status_data_start}:B{row - 1})").font = total_font
    total_paid_cell = ws.cell(row=row, column=3,
                              value=f"=SUM(C{status_data_start}:C{row - 1})")
    total_paid_cell.font = total_font
    total_paid_cell.number_format = '#,##0.00'
    row += 1

    # Blank separator
    row += 1

    # ---- Monthly Trend ------------------------------------------------------
    _bold_cell(ws, row, 1, "Monthly Trend", size=12)
    row += 1

    trend_headers = ["Month", "Claims", "Gross Cost", "Total Paid", "Paid Claims"]
    _header_row(ws, row, trend_headers)
    row += 1

    trend_data_start = row
    for _, mrow in df_monthly.iterrows():
        ws.cell(row=row, column=1, value=str(mrow["year_month"]))
        ws.cell(row=row, column=2, value=int(mrow["claim_count"]))
        ws.cell(row=row, column=3, value=float(mrow["total_gross_cost"]))
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=float(mrow["total_paid"]))
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=int(mrow["paid_count"]))
        row += 1

    # TOTAL row for Monthly Trend
    ws.cell(row=row, column=1, value="TOTAL").font = total_font
    ws.cell(row=row, column=2,
            value=f"=SUM(B{trend_data_start}:B{row - 1})").font = total_font
    gross_total = ws.cell(row=row, column=3,
                          value=f"=SUM(C{trend_data_start}:C{row - 1})")
    gross_total.font = total_font
    gross_total.number_format = '#,##0.00'
    paid_total = ws.cell(row=row, column=4,
                         value=f"=SUM(D{trend_data_start}:D{row - 1})")
    paid_total.font = total_font
    paid_total.number_format = '#,##0.00'
    ws.cell(row=row, column=5,
            value=f"=SUM(E{trend_data_start}:E{row - 1})").font = total_font
    row += 1

    # Blank separator
    row += 1

    # ---- Claim Turnaround Time -------------------------------------------------
    _bold_cell(ws, row, 1, "Claim Turnaround Time (days)", size=12)
    row += 1

    _header_row(ws, row, ["Metric", "Days"])
    row += 1

    if not df_turnaround.empty:
        r = df_turnaround.iloc[0]
        for label, val in [("Min", r["min_days"]), ("Avg", r["avg_days"]), ("Max", r["max_days"])]:
            ws.cell(row=row, column=1, value=label)
            num_cell = ws.cell(row=row, column=2, value=float(val) if val is not None else None)
            num_cell.number_format = "0.0"
            num_cell.alignment = Alignment(horizontal="right")
            row += 1

    # Column widths
    _set_col_widths(ws, [28, 16, 18, 18, 16])

    return df_monthly


def _build_detail(wb: Workbook, extra_where: str = "") -> None:
    """Build the Detail sheet with full claim-level data."""
    ws = wb.create_sheet("Detail")

    df = run_query(_inject_filter("SELECT * FROM claims ORDER BY service_date", extra_where))

    # Write header
    headers = df.columns.tolist()
    _header_row(ws, 1, headers)

    # Date format string
    date_fmt = "YYYY-MM-DD"
    currency_fmt = '"$"#,##0.00'
    date_cols = {"service_date", "paid_date"}
    cost_cols = {"gross_cost", "member_copay", "plan_paid", "total_paid"}
    numeric_cols = cost_cols | {"quantity", "days_supply", "formulary_tier"}

    # Write data rows
    for r_idx, (_, data_row) in enumerate(df.iterrows(), start=2):
        for c_idx, col_name in enumerate(headers, start=1):
            raw_val = data_row[col_name]

            # Convert to native Python types
            if pd.isna(raw_val):
                val = None
            elif col_name in date_cols:
                # Parse string date to Python date object for Excel date formatting
                try:
                    val = datetime.datetime.strptime(str(raw_val), "%Y-%m-%d").date()
                except ValueError:
                    val = str(raw_val)
            elif hasattr(raw_val, "item"):
                val = raw_val.item()
            else:
                val = raw_val

            cell = ws.cell(row=r_idx, column=c_idx, value=val)

            if col_name in date_cols:
                cell.number_format = date_fmt
            elif col_name in cost_cols:
                cell.number_format = currency_fmt

            if col_name in numeric_cols:
                cell.alignment = Alignment(horizontal="right")

    # AutoFilter covering header + all data rows
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df) + 1}"

    # Freeze top row
    ws.freeze_panes = "A2"

    # Set reasonable column widths
    col_widths = {
        "claim_id": 14, "member_id": 12, "member_name": 20, "plan_id": 10,
        "plan_name": 22, "service_date": 14, "paid_date": 14, "ndc": 14,
        "brand_name": 22, "generic_name": 22, "drug_type": 12, "therapeutic_class": 22,
        "formulary_tier": 14, "strength": 10, "days_supply": 12, "quantity": 10,
        "pharmacy_id": 14, "pharmacy_name": 22, "pharmacy_city": 16,
        "pharmacy_state": 14, "gross_cost": 14, "member_copay": 14,
        "plan_paid": 14, "total_paid": 14, "claim_status": 14,
    }
    for c_idx, col_name in enumerate(headers, start=1):
        width = col_widths.get(col_name, 15)
        ws.column_dimensions[get_column_letter(c_idx)].width = width


def _build_chart_sheet(wb: Workbook, df_monthly: pd.DataFrame, extra_where: str = "") -> None:
    """Build the Monthly Trend Chart sheet."""
    ws = wb.create_sheet("Monthly Trend Chart")

    # Write the data table that the chart references
    ws.cell(row=1, column=1, value="Month")
    ws.cell(row=1, column=2, value="Claim Count")
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2).font = Font(bold=True)

    for r_idx, (_, mrow) in enumerate(df_monthly.iterrows(), start=2):
        ws.cell(row=r_idx, column=1, value=str(mrow["year_month"]))
        ws.cell(row=r_idx, column=2, value=int(mrow["claim_count"]))

    n_months = len(df_monthly)

    date_range = _get_date_range_label(extra_where)

    # Build bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = f"Monthly Claim Volume{date_range}"
    chart.y_axis.title = "Claim Count"
    chart.x_axis.title = "Month"
    chart.style = CHART_STYLE
    chart.width = CHART_WIDTH_MD
    chart.height = CHART_HEIGHT_MD
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=n_months + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=n_months + 1)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Place chart below the data table (row n_months + 3)
    anchor_row = n_months + 3
    ws.add_chart(chart, f"A{anchor_row}")

    # Column widths for the data table
    _set_col_widths(ws, [14, 14])


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_claims_report(extra_where: str = "") -> str:
    """Build the Claims Utilization Excel report. Returns the output file path."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    queries = _load_queries(_SQL_FILE)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Summary
    df_monthly = _build_summary(wb, queries, extra_where)

    # Sheet 2: Detail
    _build_detail(wb, extra_where)

    # Sheet 3: Monthly Trend Chart
    _build_chart_sheet(wb, df_monthly, extra_where)

    # Apply print settings to all worksheets
    for ws in wb.worksheets:
        _apply_print_settings(ws)

    output_path = OUTPUT_DIR / "claims_utilization.xlsx"
    wb.save(str(output_path))

    return str(output_path)
