import pytest
from openpyxl import load_workbook
from src.reports.excel.claims_utilization import build_claims_report
from src.reports.excel.drug_cost import build_drug_report
from src.reports.excel.formulary_compliance import build_formulary_report


# Claims Utilization tests
def test_claims_detail_sheet_currency_columns_right_aligned(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    path = build_claims_report()
    wb = load_workbook(path)
    ws = wb["Detail"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    gross_col = headers.index("gross_cost") + 1
    assert ws.cell(row=2, column=gross_col).alignment.horizontal == "right"


def test_claims_detail_sheet_quantity_column_right_aligned(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    path = build_claims_report()
    wb = load_workbook(path)
    ws = wb["Detail"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    qty_col = headers.index("quantity") + 1
    assert ws.cell(row=2, column=qty_col).alignment.horizontal == "right"


# Drug Cost tests
def test_drug_cost_detail_sheet_currency_columns_right_aligned(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    path = build_drug_report()
    wb = load_workbook(path)
    ws = wb["Detail"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    gross_col = headers.index("gross_cost") + 1
    assert ws.cell(row=2, column=gross_col).alignment.horizontal == "right"


def test_drug_cost_detail_sheet_quantity_column_right_aligned(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    path = build_drug_report()
    wb = load_workbook(path)
    ws = wb["Detail"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    qty_col = headers.index("quantity") + 1
    assert ws.cell(row=2, column=qty_col).alignment.horizontal == "right"


# Formulary Compliance tests
def test_formulary_detail_sheet_currency_columns_right_aligned(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    path = build_formulary_report()
    wb = load_workbook(path)
    ws = wb["Detail"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    gross_col = headers.index("gross_cost") + 1
    assert ws.cell(row=2, column=gross_col).alignment.horizontal == "right"


def test_formulary_detail_sheet_quantity_column_right_aligned(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    path = build_formulary_report()
    wb = load_workbook(path)
    ws = wb["Detail"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    qty_col = headers.index("quantity") + 1
    assert ws.cell(row=2, column=qty_col).alignment.horizontal == "right"


def test_claims_report_all_sheets_have_print_settings(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    path = build_claims_report()
    wb = load_workbook(path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert ws.page_setup.paperSize == 1, f"{sheet_name}: expected Letter (1)"
        assert ws.page_setup.orientation == "landscape", f"{sheet_name}: expected landscape"


def test_claims_summary_has_turnaround_section(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    path = build_claims_report()
    wb = load_workbook(path)
    ws = wb["Summary"]
    cell_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any("Turnaround" in str(v) for v in cell_values if v)


def test_drug_report_has_generic_penetration_sheet(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    path = build_drug_report()
    wb = load_workbook(path)
    assert "Generic Penetration" in wb.sheetnames
    ws = wb["Generic Penetration"]
    assert ws.max_row > 1, "Generic Penetration sheet is empty"


def test_generic_penetration_headers(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    path = build_drug_report()
    wb = load_workbook(path)
    ws = wb["Generic Penetration"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
    assert headers == ["Therapeutic Class", "Total Fills", "Generic Fills", "Generic %"]
